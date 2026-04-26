"""Fast inference module without mandatory pandas import.

대형 검색어 파일에서 속도를 높이기 위해 다음을 적용한다.
- xlsx/csv 직접 읽기
- 룰은 base_rules.csv 우선 로딩
- query_norm 기준 중복 제거 후 unique query만 추론하고 원본 행에 매핑
- pandas는 선택 사항이며 기본 추론 경로에서는 import하지 않는다.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional
import csv

from .normalize import normalize_query
from .rules import load_rules, resolve_query

OUTPUT_COLUMNS = [
    "row_id", "query", "query_norm", "gate_type", "insurance_category", "customer_need_type",
    "evidence_focus", "customer_need_detail", "evidence_trace", "confidence_flag", "model_hint", "review_flag",
]


class SimpleTable:
    """Colab/CLI에서 DataFrame 없이도 결과를 다룰 수 있는 가벼운 테이블.

    - rows: list[dict]
    - columns: 컬럼명 목록
    - head(n): 앞 n개 행 SimpleTable
    - to_records(): list[dict]
    - to_dataframe(): 사용자가 명시적으로 부르면 pandas DataFrame 변환
    """

    def __init__(self, rows: List[Dict[str, Any]], columns: Optional[List[str]] = None):
        self.rows = rows
        self.columns = columns or (list(rows[0].keys()) if rows else [])

    def __len__(self) -> int:
        return len(self.rows)

    def __iter__(self):
        return iter(self.rows)

    def __getitem__(self, item):
        if isinstance(item, int):
            return self.rows[item]
        if isinstance(item, str):
            return [r.get(item, "") for r in self.rows]
        raise TypeError(item)

    def head(self, n: int = 5) -> "SimpleTable":
        return SimpleTable(self.rows[:n], self.columns)

    def to_records(self) -> List[Dict[str, Any]]:
        return list(self.rows)

    def to_dataframe(self):
        import pandas as pd
        return pd.DataFrame(self.rows, columns=self.columns)

    def to_string(self, max_rows: int = 20) -> str:
        rows = self.rows[:max_rows]
        if not rows:
            return "<empty>"
        widths = {c: min(max(len(str(c)), max(len(str(r.get(c, ""))) for r in rows)), 40) for c in self.columns}
        header = " | ".join(str(c)[:widths[c]].ljust(widths[c]) for c in self.columns)
        sep = "-+-".join("-" * widths[c] for c in self.columns)
        lines = [header, sep]
        for r in rows:
            lines.append(" | ".join(str(r.get(c, ""))[:widths[c]].ljust(widths[c]) for c in self.columns))
        if len(self.rows) > max_rows:
            lines.append(f"... ({len(self.rows)-max_rows} more rows)")
        return "\n".join(lines)

    def __repr__(self) -> str:
        return self.to_string(10)


def _read_xlsx_rows(path: Path) -> List[Dict[str, str]]:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(c).strip() if c is not None else "" for c in next(rows_iter)]
    except StopIteration:
        return []
    out: List[Dict[str, str]] = []
    for raw in rows_iter:
        item: Dict[str, str] = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            val = raw[idx] if idx < len(raw) else ""
            item[header] = "" if val is None else str(val)
        if any(str(v).strip() for v in item.values()):
            out.append(item)
    return out


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [{k: ("" if v is None else str(v)) for k, v in row.items()} for row in reader]


def _read_query_file(query_path: str | Path) -> List[Dict[str, str]]:
    path = Path(query_path)
    if not path.exists():
        raise FileNotFoundError(f"query file not found: {path}")
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        rows = _read_xlsx_rows(path)
    elif path.suffix.lower() == ".csv":
        rows = _read_csv_rows(path)
    else:
        raise ValueError(f"unsupported query file type: {path.suffix}")
    if not rows:
        return []
    if "query" not in rows[0]:
        first_col = next(iter(rows[0].keys()))
        for r in rows:
            r["query"] = r.get(first_col, "")
    return rows


def _write_xlsx(rows: List[Dict[str, Any]], columns: List[str], path: Path, sheet_name: str = "result") -> None:
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)
    for r in rows:
        ws.append([r.get(c, "") for c in columns])
    wb.save(path)


def _write_csv(rows: List[Dict[str, Any]], columns: List[str], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=columns, extrasaction="ignore")
        writer.writeheader()
        for r in rows:
            writer.writerow(r)


def _write_output(rows: List[Dict[str, Any]], columns: List[str], output_path: str | Path | None) -> None:
    if not output_path:
        return
    path = Path(output_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    if path.suffix.lower() == ".csv":
        _write_csv(rows, columns, path)
    else:
        _write_xlsx(rows, columns, path)


def _predict_hint_lazy(model_bundle: Any, query: str) -> str:
    if model_bundle is None:
        return ""
    from .model import predict_model_hint
    return predict_model_hint(model_bundle, query)


def run_inference(
    query_path: str | Path,
    base_rules_path: str | Path = "configs/base_rules.xlsx",
    feedback_store_path: str | Path | None = None,
    model_path: str | Path | None = None,
    output_path: str | Path | None = "outputs/inference_result.xlsx",
    enable_semantic_hints: bool = False,
    dedupe: bool = True,
) -> SimpleTable:
    """검색어 파일 전체 추론.

    속도 개선 핵심은 dedupe=True다. 동일 query_norm이 반복되면 한 번만 추론하고 결과를 재사용한다.
    """
    query_rows = _read_query_file(query_path)
    rulebook = load_rules(base_rules_path, feedback_store_path)
    model_bundle = None
    if model_path:
        from .model import load_model_bundle
        model_bundle = load_model_bundle(model_path)

    cache: Dict[str, Dict[str, Any]] = {}
    out_rows: List[Dict[str, Any]] = []
    extra_cols = [c for c in (list(query_rows[0].keys()) if query_rows else []) if c != "query"]
    for idx, row in enumerate(query_rows):
        query = str(row.get("query", ""))
        qn = normalize_query(query)
        if dedupe and qn in cache:
            result = dict(cache[qn])
            result["query"] = query
        else:
            hint = _predict_hint_lazy(model_bundle, query)
            result = resolve_query(query, rulebook, model_hint=hint, enable_semantic_hints=enable_semantic_hints)
            if dedupe:
                cache[qn] = dict(result)
        result = dict(result)
        result["row_id"] = idx
        for c in extra_cols:
            result[c] = row.get(c, "")
        out_rows.append(result)

    columns = OUTPUT_COLUMNS + [c for c in extra_cols if c not in OUTPUT_COLUMNS]
    _write_output(out_rows, columns, output_path)
    return SimpleTable(out_rows, columns)


def export_review_template(result_table: Any, output_path: str | Path = "outputs/review_template.xlsx") -> SimpleTable:
    rows = result_table.to_records() if hasattr(result_table, "to_records") else list(result_table)
    review_rows: List[Dict[str, Any]] = []
    for r in rows:
        item = {
            "row_id": r.get("row_id", ""),
            "query": r.get("query", ""),
            "query_norm": r.get("query_norm", ""),
            "pred_gate_type": r.get("gate_type", ""),
            "pred_insurance_category": r.get("insurance_category", ""),
            "pred_customer_need_type": r.get("customer_need_type", ""),
            "pred_evidence_focus": r.get("evidence_focus", ""),
            "customer_need_detail": r.get("customer_need_detail", ""),
            "evidence_trace": r.get("evidence_trace", ""),
            "confidence_flag": r.get("confidence_flag", ""),
            "model_hint": r.get("model_hint", ""),
            "review_flag": r.get("review_flag", ""),
            "corrected_gate_type": "",
            "corrected_insurance_category": "",
            "corrected_customer_need_type": "",
            "corrected_evidence_focus": "",
            "review_status": "",
            "memo": "",
        }
        review_rows.append(item)
    cols = list(review_rows[0].keys()) if review_rows else [
        "row_id", "query", "query_norm", "pred_gate_type", "pred_insurance_category", "pred_customer_need_type", "pred_evidence_focus",
        "customer_need_detail",
        "evidence_trace", "confidence_flag", "model_hint", "review_flag", "corrected_gate_type", "corrected_insurance_category",
        "corrected_customer_need_type", "corrected_evidence_focus", "review_status", "memo"
    ]
    _write_output(review_rows, cols, output_path)
    return SimpleTable(review_rows, cols)
