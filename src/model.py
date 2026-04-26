"""가벼운 보조 모델.

v5 Phase 1에서 모델은 최종 판단자가 아니다.
모델은 model_hint만 제공한다. 룰/Resolver가 명확하면 룰이 이긴다.

권장 모델:
- TF-IDF char/word ngram
- LogisticRegression

주의:
- evidence_focus는 절대 학습하지 않는다.
- 모델은 query에 없는 브랜드/보장/연령/상품조건을 생성하지 않는다.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any, Dict, Optional

import joblib
import pandas as pd

from .normalize import normalize_query


def _read_xlsx_as_dataframe(path: Path) -> pd.DataFrame:
    """openpyxl로 첫 번째 시트를 읽는다. 모델 학습용 보조 함수."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()
    headers = [str(c).strip() if c is not None else "" for c in rows[0]]
    data = []
    for raw in rows[1:]:
        item = {}
        for idx, header in enumerate(headers):
            if not header:
                continue
            value = raw[idx] if idx < len(raw) else ""
            item[header] = "" if value is None else str(value)
        if any(str(v).strip() for v in item.values()):
            data.append(item)
    return pd.DataFrame(data).fillna("")


def train_model(
    gold_path: str | Path,
    output_path: str | Path = "outputs/model_bundle.joblib",
) -> Dict[str, Any]:
    """gold 파일로 gate/category/need 보조 모델을 학습한다."""
    from sklearn.feature_extraction.text import TfidfVectorizer
    from sklearn.linear_model import LogisticRegression
    from sklearn.pipeline import Pipeline

    path = Path(gold_path)
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        df = _read_xlsx_as_dataframe(path).fillna("")
    else:
        df = pd.read_csv(path, dtype=str).fillna("")

    if "query" not in df.columns:
        raise ValueError("gold file must contain query column")

    df["query_norm"] = df["query"].map(normalize_query)

    targets = {}
    for col in ["gate_type", "insurance_category", "customer_need_type"]:
        if col not in df.columns:
            continue
        train_df = df[(df["query_norm"] != "") & (df[col].astype(str).str.strip() != "")]
        if train_df[col].nunique() < 2:
            continue
        pipe = Pipeline(
            steps=[
                ("tfidf", TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 5), min_df=1)),
                ("clf", LogisticRegression(max_iter=1000, class_weight="balanced")),
            ]
        )
        pipe.fit(train_df["query_norm"], train_df[col])
        targets[col] = pipe

    bundle = {"targets": targets}
    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    joblib.dump(bundle, output_path)
    return bundle


def load_model_bundle(model_path: str | Path | None) -> Optional[Dict[str, Any]]:
    if not model_path:
        return None
    path = Path(model_path)
    if not path.exists():
        return None
    return joblib.load(path)


def _predict_one(pipe: Any, query_norm: str) -> tuple[str, float]:
    pred = pipe.predict([query_norm])[0]
    prob = 0.0
    if hasattr(pipe[-1], "predict_proba"):
        probs = pipe.predict_proba([query_norm])[0]
        prob = float(max(probs))
    return str(pred), prob


def predict_model_hint(model_bundle: Optional[Dict[str, Any]], query: str) -> str:
    """모델 보조 제안을 문자열로 반환한다."""
    if not model_bundle or not model_bundle.get("targets"):
        return ""
    query_norm = normalize_query(query)
    parts = []
    name_map = {
        "gate_type": "gate",
        "insurance_category": "category",
        "customer_need_type": "need",
    }
    for target, pipe in model_bundle["targets"].items():
        value, prob = _predict_one(pipe, query_norm)
        parts.append(f"{name_map.get(target, target)}={value}|{prob:.2f}")
    return "; ".join(parts)
