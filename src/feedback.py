"""Colab Quick Fix feedback loop without mandatory pandas.

피드백은 feedback_store.xlsx 한 파일에 저장된다.
- exact_overrides: 특정 query_norm 정답 고정
- phrase_additions: 새 phrase 추가
- exclusions: 특정 slot 오탐 차단

이 모듈은 pandas 없이 openpyxl만 사용한다. Colab에서 속도와 안정성을 높이기 위한 조치다.
"""

from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List

from .normalize import normalize_label, normalize_query
from .infer import SimpleTable

EXACT_COLUMNS = ["query_norm", "gate_type", "insurance_category", "customer_need_type", "evidence_focus", "memo", "updated_at"]
PHRASE_COLUMNS = [
    "surface", "slot", "canonical_value", "level", "priority", "category_hint", "need_hint", "gate_hint",
    "is_protected", "allow_nested", "match_type", "source", "memo", "use_typo", "typo_threshold",
    "use_embedding", "embedding_text", "embedding_threshold",
]
EXCLUSION_COLUMNS = ["surface", "block_slot", "reason", "updated_at"]


def _now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def _empty_store() -> Dict[str, List[Dict[str, Any]]]:
    return {"exact_overrides": [], "phrase_additions": [], "exclusions": []}


def _read_sheet(path: Path, sheet_name: str) -> List[Dict[str, str]]:
    if not path.exists():
        return []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(c).strip() if c is not None else "" for c in next(rows)]
        except StopIteration:
            return []
        out = []
        for raw in rows:
            item = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                val = raw[idx] if idx < len(raw) else ""
                item[header] = "" if val is None else str(val)
            if any(str(v).strip() for v in item.values()):
                out.append(item)
        return out
    except Exception:
        return []


def load_feedback_store(feedback_store_path: str | Path) -> Dict[str, List[Dict[str, Any]]]:
    path = Path(feedback_store_path)
    store = _empty_store()
    if not path.exists():
        return store
    store["exact_overrides"] = _read_sheet(path, "exact_overrides")
    store["phrase_additions"] = _read_sheet(path, "phrase_additions")
    store["exclusions"] = _read_sheet(path, "exclusions")
    return store


def save_feedback_store(feedback_store_path: str | Path, store: Dict[str, List[Dict[str, Any]]]) -> None:
    from openpyxl import Workbook
    path = Path(feedback_store_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    wb.remove(wb.active)
    for sheet, columns in {
        "exact_overrides": EXACT_COLUMNS,
        "phrase_additions": PHRASE_COLUMNS,
        "exclusions": EXCLUSION_COLUMNS,
    }.items():
        ws = wb.create_sheet(sheet)
        ws.append(columns)
        for row in store.get(sheet, []):
            ws.append([row.get(c, "") for c in columns])
    wb.save(path)


def _records(table: Any) -> List[Dict[str, Any]]:
    """여러 테이블 형태를 list[dict]로 안전하게 변환한다.

    pandas DataFrame도 ``to_records`` 메서드를 갖지만, 그것을 먼저 쓰면 numpy record가 반환되어
    ``row.get(...)`` 호출에서 오류가 난다. 따라서 pandas 계열은 ``to_dict('records')`` 또는
    ``iterrows``를 먼저 사용한다.
    """
    if table is None:
        return []
    if hasattr(table, "to_dict"):
        try:
            records = table.to_dict("records")
            return [dict(x) for x in records]
        except Exception:
            pass
    if isinstance(table, list):
        return [dict(x) if hasattr(x, "items") else x for x in table]
    if hasattr(table, "iterrows"):
        return [dict(row) for _, row in table.iterrows()]
    if hasattr(table, "to_records"):
        try:
            return [dict(x) if hasattr(x, "items") else {k: x[k] for k in getattr(x, "dtype", {}).names} for x in table.to_records()]
        except Exception:
            return []
    return []


def _get_query_from_row(result_table: Any, row_id: int) -> str:
    rows = _records(result_table)
    for row in rows:
        try:
            if int(row.get("row_id", -1)) == int(row_id):
                return str(row.get("query", ""))
        except Exception:
            pass
    if 0 <= int(row_id) < len(rows):
        return str(rows[int(row_id)].get("query", ""))
    raise IndexError(f"row_id not found: {row_id}")


def _dedupe(rows: List[Dict[str, Any]], keys: List[str]) -> List[Dict[str, Any]]:
    m: Dict[tuple, Dict[str, Any]] = {}
    for row in rows:
        key = tuple(str(row.get(k, "")) for k in keys)
        m[key] = row
    return list(m.values())


def quick_fix_exact(
    feedback_store_path: str | Path,
    result_df: Any,
    row_id: int,
    gate_type: str,
    insurance_category: str,
    customer_need_type: str,
    evidence_focus: str = "null",
    memo: str = "",
) -> SimpleTable:
    query = _get_query_from_row(result_df, row_id)
    row = {
        "query_norm": normalize_query(query),
        "gate_type": normalize_label(gate_type, null_value="general"),
        "insurance_category": normalize_label(insurance_category, null_value="null"),
        "customer_need_type": normalize_label(customer_need_type, null_value="상품추천탐색"),
        "evidence_focus": normalize_label(evidence_focus, null_value="null"),
        "memo": memo,
        "updated_at": _now(),
    }
    store = load_feedback_store(feedback_store_path)
    store["exact_overrides"] = _dedupe(store["exact_overrides"] + [row], ["query_norm"])
    save_feedback_store(feedback_store_path, store)
    return SimpleTable(store["exact_overrides"], EXACT_COLUMNS)


def quick_add_phrase(
    feedback_store_path: str | Path,
    surface: str,
    slot: str,
    canonical_value: str,
    level: int,
    priority: int,
    category_hint: str = "",
    need_hint: str = "",
    gate_hint: str = "",
    is_protected: str = "N",
    allow_nested: str = "N",
    match_type: str = "contains",
    memo: str = "",
    use_typo: str = "Y",
    typo_threshold: str = "",
    use_embedding: str = "N",
    embedding_text: str = "",
    embedding_threshold: str = "",
) -> SimpleTable:
    row = {
        "surface": normalize_query(surface),
        "slot": normalize_label(slot, null_value=""),
        "canonical_value": normalize_label(canonical_value, null_value=""),
        "level": str(int(level)),
        "priority": str(int(priority)),
        "category_hint": normalize_label(category_hint, null_value=""),
        "need_hint": normalize_label(need_hint, null_value=""),
        "gate_hint": normalize_label(gate_hint, null_value=""),
        "is_protected": is_protected,
        "allow_nested": allow_nested,
        "match_type": match_type,
        "source": "feedback",
        "memo": memo,
        "use_typo": use_typo,
        "typo_threshold": typo_threshold,
        "use_embedding": use_embedding,
        "embedding_text": embedding_text,
        "embedding_threshold": embedding_threshold,
    }
    store = load_feedback_store(feedback_store_path)
    store["phrase_additions"] = _dedupe(store["phrase_additions"] + [row], ["surface", "slot", "canonical_value"])
    save_feedback_store(feedback_store_path, store)
    return SimpleTable(store["phrase_additions"], PHRASE_COLUMNS)


def quick_add_protected_span(
    feedback_store_path: str | Path,
    surface: str,
    slot: str,
    canonical_value: str,
    level: int,
    priority: int,
    category_hint: str = "",
    need_hint: str = "",
    gate_hint: str = "",
    memo: str = "",
) -> SimpleTable:
    return quick_add_phrase(
        feedback_store_path=feedback_store_path,
        surface=surface,
        slot=slot,
        canonical_value=canonical_value,
        level=level,
        priority=priority,
        category_hint=category_hint,
        need_hint=need_hint,
        gate_hint=gate_hint,
        is_protected="Y",
        allow_nested="N",
        memo=memo,
    )


def quick_add_exclusion(
    feedback_store_path: str | Path,
    surface: str,
    block_slot: str,
    reason: str = "",
) -> SimpleTable:
    row = {
        "surface": normalize_query(surface),
        "block_slot": normalize_label(block_slot, null_value=""),
        "reason": reason,
        "updated_at": _now(),
    }
    store = load_feedback_store(feedback_store_path)
    store["exclusions"] = _dedupe(store["exclusions"] + [row], ["surface", "block_slot"])
    save_feedback_store(feedback_store_path, store)
    return SimpleTable(store["exclusions"], EXCLUSION_COLUMNS)


def compile_review_template(review_path: str | Path, feedback_store_path: str | Path) -> Dict[str, int]:
    """review_template.xlsx의 approved 수정값을 exact_overrides로 반영한다."""
    rows = _read_sheet(Path(review_path), "result") or _read_sheet(Path(review_path), "review") or _read_sheet(Path(review_path), "Sheet1")
    store = load_feedback_store(feedback_store_path)
    added = 0
    for r in rows:
        if str(r.get("review_status", "")).strip().lower() != "approved":
            continue
        query_norm = normalize_query(r.get("query_norm", "") or r.get("query", ""))
        if not query_norm:
            continue
        row = {
            "query_norm": query_norm,
            "gate_type": normalize_label(r.get("corrected_gate_type", r.get("pred_gate_type", "")), null_value="general"),
            "insurance_category": normalize_label(r.get("corrected_insurance_category", r.get("pred_insurance_category", "")), null_value="null"),
            "customer_need_type": normalize_label(r.get("corrected_customer_need_type", r.get("pred_customer_need_type", "")), null_value="상품추천탐색"),
            "evidence_focus": normalize_label(r.get("corrected_evidence_focus", r.get("pred_evidence_focus", "null")), null_value="null"),
            "memo": str(r.get("memo", "") or "review_template"),
            "updated_at": _now(),
        }
        store["exact_overrides"].append(row)
        added += 1
    store["exact_overrides"] = _dedupe(store["exact_overrides"], ["query_norm"])
    save_feedback_store(feedback_store_path, store)
    return {"added_exact_overrides": added, "total_exact_overrides": len(store["exact_overrides"])}
