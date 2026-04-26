"""Fast rule engine for v5 Phase 1 Light.

이 파일은 검색어 이해 엔진의 핵심이다. 이전 버전은 pandas DataFrame + 전체 룰 선형 스캔에
가깝게 동작했기 때문에 대형 검색어 파일에서 느릴 수 있었다. 이 버전은 다음을 보강한다.

1. pandas 의존 제거
   - 룰 로딩은 csv/openpyxl 표준 방식으로 수행한다.
   - Colab 환경에서 pandas import/엑셀 엔진 이슈로 멈추는 문제를 줄인다.

2. CSV 우선 로딩
   - base_rules.xlsx가 들어와도 같은 폴더에 base_rules.csv가 있으면 csv를 우선 사용한다.
   - 대형 룰 사전은 xlsx보다 csv 로딩이 훨씬 빠르다.

3. 룰 인덱싱
   - 모든 룰을 매번 전수 검사하지 않고, compact surface의 2-gram index로 후보 룰만 좁힌다.
   - 띄어쓰기 흔들림도 compact_contains로 잡는다.
     예: '라이나 치아 보험' -> compact '라이나치아보험'에서 '치아보험' 매칭.

4. Similarity Matcher 경량화
   - 오탈자 보정은 brand/category/coverage/underwriting에 한정한다.
   - 의미 유사 힌트는 기본 OFF다. 애매한 의미 추론은 답을 만들지 않고 review로 보낸다.

5. Resolver 계약 유지
   - 최종 판단은 Semantic Resolver LEVEL 우선순위가 담당한다.
   - OOS/Public은 similarity가 뒤집을 수 없다.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from difflib import SequenceMatcher
from pathlib import Path
import csv
import math
import re
from typing import Any, Dict, Iterable, List, Optional, Tuple

from .normalize import normalize_label, normalize_query
from .schema import ALLOWED_CATEGORIES, ALLOWED_GATE_TYPES, ALLOWED_NEEDS


LEVELS = {
    "exact_override": 0,
    "hard_oos": 1,
    "public_non_private": 2,
    "action_service": 3,
    "policy_product_feature": 4,
    "underwriting_type": 4,
    "coverage_focus": 5,
    "disease_focus": 5,
    "target_segment": 5,
    "insurance_category": 6,
    "brand_name": 7,
    "weak_intent": 8,
    "insurance_intent": 8,
    "definition_signal": 8,
    "fallback": 9,
}

TYPO_ALLOWED_SLOTS = {"brand_name", "insurance_category", "coverage_focus", "underwriting_type"}
SEMANTIC_HINT_SLOTS = {"coverage_focus", "policy_product_feature", "underwriting_type", "insurance_category"}

PHRASE_COLUMNS = [
    "surface", "slot", "canonical_value", "level", "priority", "category_hint", "need_hint", "gate_hint",
    "is_protected", "allow_nested", "match_type", "source", "memo", "use_typo", "typo_threshold",
    "use_embedding", "embedding_text", "embedding_threshold",
]
EXACT_COLUMNS = ["query_norm", "gate_type", "insurance_category", "customer_need_type", "evidence_focus", "memo", "updated_at"]
EXCLUSION_COLUMNS = ["surface", "block_slot", "reason", "updated_at"]


@dataclass(frozen=True)
class Rule:
    surface: str
    surface_compact: str
    slot: str
    canonical_value: str
    level: int
    priority: int
    category_hint: str = ""
    need_hint: str = ""
    gate_hint: str = ""
    is_protected: bool = False
    allow_nested: bool = False
    match_type: str = "contains"
    source: str = "base"
    memo: str = ""
    use_typo: bool = False
    typo_threshold: float = 0.80
    use_embedding: bool = False
    embedding_text: str = ""
    embedding_threshold: float = 0.82
    embedding_vec: Dict[str, float] = field(default_factory=dict)


@dataclass
class Span:
    surface: str
    slot: str
    canonical_value: str
    start: int
    end: int
    level: int
    priority: int
    category_hint: str = ""
    need_hint: str = ""
    gate_hint: str = ""
    is_protected: bool = False
    allow_nested: bool = False
    match_type: str = "contains"
    source: str = "base"
    match_score: float = 0.90
    hint_only: bool = False

    @property
    def length(self) -> int:
        return max(0, self.end - self.start)

    def overlaps(self, other: "Span") -> bool:
        return self.start < other.end and other.start < self.end

    def inside(self, other: "Span") -> bool:
        return self.start >= other.start and self.end <= other.end


@dataclass
class RuleBook:
    phrases: Any = field(default_factory=list)
    exact_overrides: Any = field(default_factory=list)
    exclusions: Any = field(default_factory=list)
    indexed_rules: Dict[str, List[Rule]] = field(default_factory=dict)
    short_rules: List[Rule] = field(default_factory=list)
    regex_rules: List[Rule] = field(default_factory=list)
    typo_rules_by_len: Dict[int, List[Rule]] = field(default_factory=dict)
    typo_index: Dict[str, List[Rule]] = field(default_factory=dict)
    semantic_rules: List[Rule] = field(default_factory=list)

    def __post_init__(self) -> None:
        # 테스트/하위호환: pandas DataFrame이 직접 들어와도 list[dict]로 변환한다.
        phrase_rows = _rows_any(self.phrases)
        exact_rows = _rows_any(self.exact_overrides)
        exclusion_rows = _rows_any(self.exclusions)
        self.phrases = [_standardize_rule(r, default_source=str(r.get("source", "test"))) for r in phrase_rows]
        self.phrases = [r for r in self.phrases if r.surface and r.slot]
        self.exact_overrides = [_standardize_exact_row(r) for r in exact_rows if normalize_query(r.get("query_norm", ""))]
        self.exclusions = [_standardize_exclusion_row(r) for r in exclusion_rows if normalize_query(r.get("surface", "")) and normalize_label(r.get("block_slot", ""), null_value="")]
        self._build_indexes()

    def _build_indexes(self) -> None:
        self.indexed_rules = {}
        self.short_rules = []
        self.regex_rules = []
        self.typo_rules_by_len = {}
        self.typo_index = {}
        self.semantic_rules = []
        seen: set[Tuple[str, str, str]] = set()
        deduped: List[Rule] = []
        for r in self.phrases:
            key = (r.surface, r.slot, r.canonical_value)
            if key in seen:
                continue
            seen.add(key)
            deduped.append(r)
        self.phrases = deduped
        for rule in self.phrases:
            if rule.match_type == "regex":
                self.regex_rules.append(rule)
            elif len(rule.surface_compact) < 2:
                self.short_rules.append(rule)
            else:
                grams = _bigrams(rule.surface_compact)
                key = min(grams, key=len) if grams else rule.surface_compact[:2]
                self.indexed_rules.setdefault(key, []).append(rule)
            if rule.use_typo and rule.slot in TYPO_ALLOWED_SLOTS and len(rule.surface_compact) >= 2:
                self.typo_rules_by_len.setdefault(len(rule.surface_compact), []).append(rule)
                for gram in _bigrams(rule.surface_compact):
                    self.typo_index.setdefault(gram, []).append(rule)
            if rule.use_embedding and rule.slot in SEMANTIC_HINT_SLOTS:
                self.semantic_rules.append(rule)
        for rules in self.indexed_rules.values():
            rules.sort(key=lambda r: (r.level, -r.priority, -len(r.surface_compact), r.surface))
        self.regex_rules.sort(key=lambda r: (r.level, -r.priority, -len(r.surface_compact)))
        self.short_rules.sort(key=lambda r: (r.level, -r.priority, -len(r.surface_compact)))


_RULEBOOK_CACHE: Dict[Tuple[str, str, float, float], RuleBook] = {}


def _rows_any(obj: Any) -> List[Dict[str, Any]]:
    if obj is None:
        return []
    if isinstance(obj, list):
        return [dict(x) for x in obj]
    if isinstance(obj, tuple):
        return [dict(x) for x in obj]
    if hasattr(obj, "iterrows"):
        return [dict(row) for _, row in obj.iterrows()]
    return []


def _clean_bool(value: Any) -> bool:
    return str(value).strip().upper() in {"Y", "YES", "TRUE", "1"}


def _to_float(value: Any, default: float) -> float:
    try:
        text = str(value).strip()
        return float(text) if text else default
    except Exception:
        return default


def _to_int(value: Any, default: int) -> int:
    try:
        text = str(value).strip()
        return int(float(text)) if text else default
    except Exception:
        return default


def _default_typo_threshold(slot: str) -> float:
    if slot == "brand_name":
        return 0.82
    if slot == "insurance_category":
        return 0.88
    if slot == "coverage_focus":
        return 0.90
    return 0.82


def _default_embedding_text(row: Dict[str, Any]) -> str:
    pieces = [row.get("surface", ""), row.get("canonical_value", ""), row.get("category_hint", ""), row.get("need_hint", ""), row.get("memo", "")]
    return normalize_query(" ".join(str(p) for p in pieces if str(p).strip()))


def _standardize_rule(row: Dict[str, Any], default_source: str = "base") -> Rule:
    surface = normalize_query(row.get("surface", ""))
    slot = normalize_label(row.get("slot", ""), null_value="")
    canonical = normalize_label(row.get("canonical_value", surface), null_value="") or surface
    level = _to_int(row.get("level", ""), LEVELS.get(slot, 9))
    priority = _to_int(row.get("priority", ""), 0)
    use_typo = _clean_bool(row.get("use_typo", "")) or slot in TYPO_ALLOWED_SLOTS
    # policy_product_feature는 오탐이 커서 typo 기본값 제외한다.
    if slot == "policy_product_feature":
        use_typo = _clean_bool(row.get("use_typo", ""))
    # underwriting_type 숫자 룰(225/325/355 등)은 2025 같은 연도/가격에서 오탈자 후보로 잡히면 치명적이다.
    # 따라서 명시적으로 Y가 들어온 경우를 제외하고 숫자형 underwriting 룰의 typo는 끈다.
    surface_compact_for_typo = surface.replace(" ", "")
    if slot == "underwriting_type" and re.fullmatch(r"\d{2,5}", surface_compact_for_typo):
        use_typo = _clean_bool(row.get("use_typo", "")) and False
    use_embedding = _clean_bool(row.get("use_embedding", "")) and slot in SEMANTIC_HINT_SLOTS
    emb_text = normalize_query(row.get("embedding_text", "")) or _default_embedding_text(row)
    return Rule(
        surface=surface,
        surface_compact=surface.replace(" ", ""),
        slot=slot,
        canonical_value=canonical,
        level=level,
        priority=priority,
        category_hint=normalize_label(row.get("category_hint", ""), null_value=""),
        need_hint=normalize_label(row.get("need_hint", ""), null_value=""),
        gate_hint=normalize_label(row.get("gate_hint", ""), null_value=""),
        is_protected=_clean_bool(row.get("is_protected", "")),
        allow_nested=_clean_bool(row.get("allow_nested", "")),
        match_type=normalize_label(row.get("match_type", "contains"), null_value="contains"),
        source=str(row.get("source", default_source) or default_source),
        memo=str(row.get("memo", "") or ""),
        use_typo=use_typo,
        typo_threshold=_to_float(row.get("typo_threshold", ""), _default_typo_threshold(slot)),
        use_embedding=use_embedding,
        embedding_text=emb_text,
        embedding_threshold=_to_float(row.get("embedding_threshold", ""), 0.82),
        embedding_vec=_char_ngrams(emb_text) if use_embedding else {},
    )


def _standardize_exact_row(row: Dict[str, Any]) -> Dict[str, str]:
    return {
        "query_norm": normalize_query(row.get("query_norm", "")),
        "gate_type": normalize_label(row.get("gate_type", ""), null_value=""),
        "insurance_category": normalize_label(row.get("insurance_category", ""), null_value="null"),
        "customer_need_type": normalize_label(row.get("customer_need_type", ""), null_value=""),
        "evidence_focus": normalize_label(row.get("evidence_focus", ""), null_value="null"),
        "memo": str(row.get("memo", "") or ""),
        "updated_at": str(row.get("updated_at", "") or ""),
    }


def _standardize_exclusion_row(row: Dict[str, Any]) -> Dict[str, str]:
    return {
        "surface": normalize_query(row.get("surface", "")),
        "block_slot": normalize_label(row.get("block_slot", ""), null_value=""),
        "reason": str(row.get("reason", "") or ""),
        "updated_at": str(row.get("updated_at", "") or ""),
    }


def _path_cache_key(path: str | Path | None) -> Tuple[str, float]:
    if not path:
        return "", 0.0
    p = Path(path)
    if not p.exists():
        return str(p.resolve()), 0.0
    return str(p.resolve()), p.stat().st_mtime


def _prefer_csv(path: Path) -> Path:
    """xlsx 대신 같은 이름의 csv가 있으면 csv를 우선 사용한다."""
    if path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        csv_path = path.with_suffix(".csv")
        if csv_path.exists():
            return csv_path
    return path


def _read_csv_rows(path: Path) -> List[Dict[str, str]]:
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        return [{k: ("" if v is None else str(v)) for k, v in row.items()} for row in reader]


def _read_xlsx_rows(path: Path, sheet_name: str) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        if sheet_name not in wb.sheetnames:
            return []
        ws = wb[sheet_name]
        rows = ws.iter_rows(values_only=True)
        try:
            headers = [str(c).strip() if c is not None else "" for c in next(rows)]
        except StopIteration:
            return []
        out = []
        for raw in rows:
            item = {}
            for idx, header in enumerate(headers):
                if not header:
                    continue
                val = raw[idx] if idx < len(raw) else ""
                item[header] = "" if val is None else str(val)
            if any(str(v).strip() for v in item.values()):
                out.append(item)
        return out
    except Exception:
        return []


def _read_table(path: str | Path, sheet_name: str) -> List[Dict[str, str]]:
    p = Path(path)
    if not p.exists():
        return []
    if sheet_name == "phrases":
        p = _prefer_csv(p)
    if p.suffix.lower() == ".csv":
        return _read_csv_rows(p) if sheet_name == "phrases" else []
    if p.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        return _read_xlsx_rows(p, sheet_name)
    return []


def load_rules(base_rules_path: str | Path = "configs/base_rules.xlsx", feedback_store_path: str | Path | None = None) -> RuleBook:
    base_path = _prefer_csv(Path(base_rules_path))
    base_key, base_mtime = _path_cache_key(base_path)
    feedback_key, feedback_mtime = _path_cache_key(feedback_store_path)
    cache_key = (base_key, feedback_key, base_mtime, feedback_mtime)
    if cache_key in _RULEBOOK_CACHE:
        return _RULEBOOK_CACHE[cache_key]

    phrases = _read_table(base_path, "phrases")
    exclusions = _read_table(base_rules_path, "exclusions")
    exact: List[Dict[str, str]] = []
    if feedback_store_path and Path(feedback_store_path).exists():
        phrases += _read_table(feedback_store_path, "phrase_additions")
        exclusions += _read_table(feedback_store_path, "exclusions")
        exact += _read_table(feedback_store_path, "exact_overrides")
    rb = RuleBook(phrases=phrases, exact_overrides=exact, exclusions=exclusions)
    _RULEBOOK_CACHE[cache_key] = rb
    return rb


def _bigrams(text: str) -> set[str]:
    if len(text) < 2:
        return {text} if text else set()
    return {text[i:i+2] for i in range(len(text)-1)}


def _query_candidate_rules(query_norm: str, rulebook: RuleBook) -> List[Rule]:
    compact = query_norm.replace(" ", "")
    grams = _bigrams(compact)
    candidates: Dict[Tuple[str, str, str], Rule] = {}
    for gram in grams:
        for rule in rulebook.indexed_rules.get(gram, []):
            candidates[(rule.surface, rule.slot, rule.canonical_value)] = rule
    for rule in rulebook.short_rules:
        candidates[(rule.surface, rule.slot, rule.canonical_value)] = rule
    for rule in rulebook.regex_rules:
        candidates[(rule.surface, rule.slot, rule.canonical_value)] = rule
    return list(candidates.values())


def _rule_to_span(rule: Rule, surface: str, start: int, end: int, match_type: str, score: float) -> Span:
    return Span(
        surface=surface,
        slot=rule.slot,
        canonical_value=rule.canonical_value or surface,
        start=int(start),
        end=int(end),
        level=int(rule.level),
        priority=int(rule.priority),
        category_hint=rule.category_hint,
        need_hint=rule.need_hint,
        gate_hint=rule.gate_hint,
        is_protected=rule.is_protected,
        allow_nested=rule.allow_nested,
        match_type=match_type,
        source=rule.source,
        match_score=float(score),
    )


def _iter_matches(query_norm: str, rule: Rule) -> List[Tuple[int, int, str, float]]:
    surface = rule.surface
    if not surface:
        return []
    if rule.match_type == "regex":
        try:
            return [(m.start(), m.end(), "regex", 0.88) for m in re.finditer(surface, query_norm)]
        except re.error:
            return []

    # 숫자형 간편심사 룰(225/325/355 등)은 2025, 가격, 상품번호 안에서 부분문자열로 잡히면
    # 임플란트 가격 2025 → 225간편심사 같은 치명적 오탐이 발생한다.
    # 따라서 순수 숫자 underwriting_type은 좌우가 숫자가 아닐 때만 매칭한다.
    if rule.slot == "underwriting_type" and re.fullmatch(r"\d{2,5}", surface):
        pattern = rf"(?<!\d){re.escape(surface)}(?!\d)"
        return [(m.start(), m.end(), "contains", 0.90) for m in re.finditer(pattern, query_norm)]

    matches: List[Tuple[int, int, str, float]] = []
    for m in re.finditer(re.escape(surface), query_norm):
        score = 1.0 if query_norm == surface else 0.90
        matches.append((m.start(), m.start() + len(surface), "contains", score))
    # 띄어쓰기 변형 보정: '치아 보험' ↔ '치아보험', '통합 간병 보험' ↔ '통합간병보험'
    compact = query_norm.replace(" ", "")
    surface_compact = rule.surface_compact
    # 순수 숫자 underwriting_type은 compact_contains도 금지한다. 2025 안의 225 오탐 방지.
    if not (rule.slot == "underwriting_type" and re.fullmatch(r"\d{2,5}", surface_compact)):
        if len(surface_compact) >= 3 and surface not in query_norm and surface_compact in compact:
            cstart = compact.find(surface_compact)
            matches.append((cstart, cstart + len(surface_compact), "compact_contains", 0.88))
    return matches


def detect_phrases(query_norm: str, rulebook: RuleBook) -> List[Span]:
    spans: List[Span] = []
    if not query_norm:
        return spans
    for rule in _query_candidate_rules(query_norm, rulebook):
        for start, end, match_type, score in _iter_matches(query_norm, rule):
            spans.append(_rule_to_span(rule, surface=rule.surface if match_type != "compact_contains" else rule.surface, start=start, end=end, match_type=match_type, score=score))
    spans.sort(key=lambda s: (s.level, -s.priority, -s.match_score, -s.length, s.start))
    return spans


def _decompose_hangul(text: str) -> str:
    chosung = "ㄱㄲㄴㄷㄸㄹㅁㅂㅃㅅㅆㅇㅈㅉㅊㅋㅌㅍㅎ"
    jungsung = "ㅏㅐㅑㅒㅓㅔㅕㅖㅗㅘㅙㅚㅛㅜㅝㅞㅟㅠㅡㅢㅣ"
    jongsung = "_ㄱㄲㄳㄴㄵㄶㄷㄹㄺㄻㄼㄽㄾㄿㅀㅁㅂㅄㅅㅆㅇㅈㅊㅋㅌㅍㅎ"
    out: List[str] = []
    for ch in str(text):
        code = ord(ch)
        if 0xAC00 <= code <= 0xD7A3:
            idx = code - 0xAC00
            out.append(chosung[idx // 588])
            out.append(jungsung[(idx % 588) // 28])
            jong = idx % 28
            if jong:
                out.append(jongsung[jong])
        else:
            out.append(ch)
    return "".join(out)


def _similarity(a: str, b: str) -> float:
    a_norm = normalize_query(a).replace(" ", "")
    b_norm = normalize_query(b).replace(" ", "")
    if not a_norm or not b_norm:
        return 0.0
    raw = SequenceMatcher(None, a_norm, b_norm).ratio()
    jamo = SequenceMatcher(None, _decompose_hangul(a_norm), _decompose_hangul(b_norm)).ratio()
    return max(raw, jamo)


def _candidate_windows(query_norm: str) -> List[Tuple[str, int, int]]:
    windows: List[Tuple[str, int, int]] = []
    seen = set()
    for m in re.finditer(r"[^\s]+", query_norm):
        token = m.group(0)
        if token not in seen:
            seen.add(token)
            windows.append((token, m.start(), m.end()))
    compact = query_norm.replace(" ", "")
    if len(compact) <= 60:
        for size in range(2, min(8, len(compact)) + 1):
            for start in range(0, len(compact) - size + 1):
                token = compact[start:start+size]
                if token not in seen:
                    seen.add(token)
                    windows.append((token, start, start+size))
    return windows


def _char_ngrams(text: str, n_values: Tuple[int, ...] = (2, 3, 4)) -> Dict[str, float]:
    compact = normalize_query(text).replace(" ", "")
    grams: Dict[str, float] = {}
    for n in n_values:
        if len(compact) < n:
            continue
        for i in range(len(compact)-n+1):
            gram = compact[i:i+n]
            grams[gram] = grams.get(gram, 0.0) + 1.0
    return grams


def _cosine_dict(a: Dict[str, float], b: Dict[str, float]) -> float:
    if not a or not b:
        return 0.0
    dot = sum(v * b.get(k, 0.0) for k, v in a.items())
    norm_a = math.sqrt(sum(v*v for v in a.values()))
    norm_b = math.sqrt(sum(v*v for v in b.values()))
    return dot / (norm_a * norm_b) if norm_a and norm_b else 0.0


def detect_similarity_candidates(query_norm: str, rulebook: RuleBook, exact_spans: List[Span], enable_semantic_hints: bool = False) -> Tuple[List[Span], List[str]]:
    hints: List[str] = []
    typo_spans: List[Span] = []
    if not query_norm:
        return typo_spans, hints
    exact_slots = {s.slot for s in exact_spans if s.match_type in {"contains", "regex", "compact_contains"}}
    existing = {(s.slot, s.canonical_value) for s in exact_spans}
    has_hard_stop = any(s.slot in {"hard_oos", "public_non_private"} for s in exact_spans)
    if not has_hard_stop:
        windows = _candidate_windows(query_norm)
        exact_strong_slots = exact_slots & {"brand_name", "insurance_category", "coverage_focus", "underwriting_type"}
        exact_surface_compacts = [s.surface.replace(" ", "") for s in exact_spans]
        for token, start, end in windows:
            token_compact = token.replace(" ", "")
            # 이미 정확 매칭 surface에 포함된 window는 typo 후보로 비교하지 않는다.
            if any(token_compact and (token_compact in surf or surf in token_compact) for surf in exact_surface_compacts):
                continue
            token_grams = _bigrams(token_compact)
            candidate_map: Dict[Tuple[str, str, str], Rule] = {}
            for gram in token_grams:
                for rule in rulebook.typo_index.get(gram, []):
                    candidate_map[(rule.surface, rule.slot, rule.canonical_value)] = rule
            for rule in sorted(candidate_map.values(), key=lambda r: (r.slot, abs(len(r.surface_compact)-len(token_compact)), -r.priority)):
                if rule.slot in exact_slots or (rule.slot, rule.canonical_value) in existing:
                    continue
                # 이미 brand+category 등 강한 신호가 충분하면 불필요한 coverage/category typo를 계산하지 않는다.
                if {"brand_name", "insurance_category"}.issubset(exact_strong_slots) and rule.slot != "underwriting_type":
                    continue
                # 카테고리는 공백 토큰 중심으로만 typo. window로 category가 과도하게 붙는 것 방지.
                if rule.slot == "insurance_category" and token not in query_norm.split():
                    continue
                if abs(len(token_compact) - len(rule.surface_compact)) > 1:
                    continue
                if token_compact != rule.surface_compact and (token_compact in rule.surface_compact or rule.surface_compact in token_compact):
                    continue
                score = _similarity(token, rule.surface)
                if score >= rule.typo_threshold:
                    typo_spans.append(_rule_to_span(rule, surface=token, start=start, end=end, match_type="typo", score=score))
                    existing.add((rule.slot, rule.canonical_value))
                    hints.append(f"typo:{rule.slot}={rule.canonical_value}|surface={token}|score={score:.2f}")
                    break
    if enable_semantic_hints and not has_hard_stop:
        qvec = _char_ngrams(query_norm)
        best_by_slot: Dict[str, Tuple[float, Rule]] = {}
        for rule in rulebook.semantic_rules:
            if (rule.slot, rule.canonical_value) in existing:
                continue
            score = _cosine_dict(qvec, rule.embedding_vec)
            if score >= rule.embedding_threshold:
                old = best_by_slot.get(rule.slot)
                if old is None or score > old[0]:
                    best_by_slot[rule.slot] = (score, rule)
        for slot, (score, rule) in best_by_slot.items():
            hints.append(f"embedding:{slot}={rule.canonical_value}|score={score:.2f}|hint_only=true")
    return typo_spans, hints


def build_protected_spans(spans: List[Span]) -> List[Span]:
    candidates = [s for s in spans if s.is_protected or s.level in {1, 2}]
    candidates.sort(key=lambda s: (s.start, -s.length, s.level, -s.priority))
    selected: List[Span] = []
    for s in candidates:
        if not any(s.overlaps(p) for p in selected):
            selected.append(s)
    return selected


def _is_blocked_by_exclusion(span: Span, query_norm: str, exclusions: List[Dict[str, str]]) -> Optional[str]:
    for row in exclusions:
        if row.get("block_slot") == span.slot and row.get("surface") and row["surface"] in query_norm:
            return row["surface"]
    return None


def extract_evidence(query_norm: str, spans: List[Span], protected_spans: List[Span], rulebook: RuleBook) -> Tuple[List[Span], List[str]]:
    evidence: List[Span] = []
    blocked: List[str] = []
    for span in spans:
        skip = False
        for prot in protected_spans:
            same = span.start == prot.start and span.end == prot.end and span.slot == prot.slot and span.canonical_value == prot.canonical_value
            if same:
                continue
            if span.inside(prot) and not prot.allow_nested:
                blocked.append(f"{span.slot}:{span.surface} blocked_by protected:{prot.surface}")
                skip = True
                break
        if skip:
            continue
        ex = _is_blocked_by_exclusion(span, query_norm, rulebook.exclusions)
        if ex:
            blocked.append(f"{span.slot}:{span.surface} blocked_by exclusion:{ex}")
            continue
        evidence.append(span)
    unique: Dict[Tuple[int, int, str, str, str], Span] = {}
    for s in evidence:
        unique[(s.start, s.end, s.slot, s.canonical_value, s.match_type)] = s
    out = list(unique.values())
    out.sort(key=lambda s: (s.level, -s.priority, -s.match_score, -s.length, s.start))
    return out, blocked


def find_exact_override(query_norm: str, rulebook: RuleBook) -> Optional[Dict[str, str]]:
    for row in reversed(rulebook.exact_overrides):
        if row.get("query_norm") == query_norm:
            return {
                "gate_type": row.get("gate_type", ""),
                "insurance_category": row.get("insurance_category", "null"),
                "customer_need_type": row.get("customer_need_type", ""),
                "evidence_focus": row.get("evidence_focus", "null"),
                "evidence_trace": "exact_override:selected=true",
                "confidence_flag": "high",
                "model_hint": "",
                "review_flag": "N",
            }
    return None


def _has_private_insurance_signal(evidence: List[Span]) -> bool:
    strong = {"insurance_category", "coverage_focus", "policy_product_feature", "underwriting_type", "action_service", "brand_name"}
    return any(s.slot in strong for s in evidence) or any(s.slot == "insurance_intent" for s in evidence)



BENEFIT_KEYWORDS = {
    "진단비", "진단금", "수술비", "입원", "입원비", "일당", "치료비", "의료비", "통원",
    "보장", "담보", "간병비", "간병인", "약제비", "검사비", "후유장해", "사망", "치료", "처방",
}
JOIN_KEYWORDS = {
    "가입", "가입가능", "가입 가능", "가입되", "들수", "들 수", "가능", "있어도", "있는데", "병력",
    "유병", "유병자", "간편", "간편심사", "간편고지", "고지", "고지의무", "인수", "인수기준",
    "완화", "거절", "심사", "부담보", "할증", "표준체", "무심사", "초간편",
}
COMPARE_KEYWORDS = {"비교", "순위", "랭킹", "추천", "후기", "디시", "클리앙", "가격", "보험료", "견적", "상담"}


def _joined_text(evidence: List[Span]) -> str:
    return " ".join([s.surface + " " + s.canonical_value + " " + s.slot for s in evidence])


def _has_benefit_intent(evidence: List[Span]) -> bool:
    text = _joined_text(evidence)
    if any(k in text for k in BENEFIT_KEYWORDS):
        return True
    # 진단비보험, 수술비보험 등 category 자체가 benefit 상품군이면 보장 확인 맥락이다.
    return any(s.slot == "insurance_category" and any(k in s.canonical_value for k in ["진단비", "수술비", "입원비", "치료비", "간병비", "후유장해"]) for s in evidence)


def _has_join_intent(evidence: List[Span]) -> bool:
    text = _joined_text(evidence)
    if any(k in text for k in JOIN_KEYWORDS):
        return True
    return any(s.slot in {"underwriting_type"} for s in evidence)


def _has_compare_or_commercial_intent(evidence: List[Span]) -> bool:
    text = _joined_text(evidence)
    return any(k in text for k in COMPARE_KEYWORDS)


def _disease_spans(evidence: List[Span]) -> List[Span]:
    return [s for s in evidence if s.slot == "disease_focus" or (s.slot == "coverage_focus" and s.canonical_value in {"당뇨", "고혈압", "고지혈증", "고혈압합병증", "당뇨합병증"})]


def _best_benefit_span(evidence: List[Span]) -> Optional[Span]:
    benefit = []
    for s in evidence:
        if s.slot == "coverage_focus" and any(k in (s.surface + s.canonical_value) for k in BENEFIT_KEYWORDS):
            benefit.append(s)
        elif s.slot == "insurance_category" and any(k in s.canonical_value for k in ["진단비", "수술비", "입원비", "치료비", "간병비", "후유장해"]):
            # category는 focus 후보로는 약하므로 뒤에서만 쓴다.
            pass
    if not benefit:
        return None
    benefit.sort(key=lambda s: (s.level, -s.priority, -s.match_score, -s.length, s.start))
    return benefit[0]


def _disease_context(evidence: List[Span]) -> str:
    """질환 단서의 의도 맥락을 구분한다.

    - 가입/고지/간편/병력 맥락이면 가입가능성
    - 진단비/수술비/입원/치료/보장 맥락이면 보장범위
    - 둘 다 없으면 애매하므로 disease_ambiguous
    """
    if not _disease_spans(evidence):
        return "none"
    if _has_join_intent(evidence):
        return "join"
    if _has_benefit_intent(evidence):
        return "benefit"
    return "ambiguous"

def _best_category(evidence: List[Span], selected: Optional[Span]) -> str:
    cats = [s for s in evidence if s.slot == "insurance_category"]
    has_female = any(s.slot == "target_segment" and "여성" in s.canonical_value for s in evidence)
    if selected and selected.category_hint and has_female and selected.category_hint.startswith("여성"):
        return selected.category_hint

    # context override: 특정 담보 phrase가 매우 구체적인 category_hint를 갖고 있으면 일반 category보다 우선한다.
    # 예: '치아 골절 보험'은 generic '골절보험' category보다 selected coverage의 '치아보험' hint가 더 정확하다.
    # 단, 면책기간/감액기간 같은 policy term의 generic category_hint=기타보험이 실제 상품군을 덮어쓰면
    # '간병인 보험 면책기간 -> 기타보험' 같은 오탐이 생긴다. 그래서 coverage/underwriting의 구체 hint만
    # cats보다 우선시키고, policy_product_feature는 카테고리 phrase가 없을 때만 힌트를 쓴다.
    generic_hints = {"", "null", "기타보험"}
    if selected and selected.category_hint and selected.category_hint in ALLOWED_CATEGORIES:
        if selected.slot == "coverage_focus" and selected.priority >= 8050 and selected.category_hint not in generic_hints:
            return selected.category_hint
        if selected.slot == "underwriting_type" and selected.priority >= 8050 and selected.category_hint not in generic_hints:
            return selected.category_hint
        if selected.slot == "policy_product_feature" and not cats and selected.category_hint not in generic_hints:
            return selected.category_hint

    # 질환 + 진단비/수술비/입원/치료비 맥락에서는 유병자보험보다 해당 급부 상품군을 우선한다.
    ctx = _disease_context(evidence)
    if ctx == "benefit":
        benefit = _best_benefit_span(evidence)
        if benefit and benefit.category_hint:
            return benefit.category_hint
        benefit_cats = [s for s in cats if any(k in s.canonical_value for k in ["진단비", "수술비", "입원비", "치료비", "간병비", "후유장해"])]
        if benefit_cats:
            benefit_cats.sort(key=lambda s: (-s.priority, -s.match_score, -s.length, s.start))
            return benefit_cats[0].canonical_value

    if cats:
        cats.sort(key=lambda s: (-s.priority, -s.match_score, -s.length, s.start))
        return cats[0].canonical_value
    if selected and selected.category_hint:
        return selected.category_hint
    hinted = [s for s in evidence if s.category_hint]
    if hinted:
        hinted.sort(key=lambda s: (s.level, -s.priority, -s.match_score, -s.length))
        return hinted[0].category_hint
    if any(s.slot == "insurance_intent" for s in evidence):
        return "기타보험"
    return "null"


def _selected_evidence(evidence: List[Span]) -> Optional[Span]:
    candidates = [s for s in evidence if s.slot != "insurance_intent" and not s.hint_only]
    if not candidates:
        candidates = evidence[:]
    if not candidates:
        return None

    # 질환명(당뇨/고혈압/고지혈증 등)은 문맥 없이 무조건 가입가능성으로 보내지 않는다.
    # 진단비/수술비/입원/치료/보장 맥락이 있으면 benefit span을 우선 선택한다.
    ctx = _disease_context(evidence)
    if ctx == "benefit":
        benefit = _best_benefit_span(evidence)
        if benefit:
            return benefit
    elif ctx == "join":
        ds = _disease_spans(evidence)
        if ds:
            ds.sort(key=lambda s: (s.level, -s.priority, -s.match_score, -s.length, s.start))
            return ds[0]
    elif ctx == "ambiguous":
        # 질환명 + 보험만 있는 경우는 유병자 가입가능성으로 확정하지 않고 질환 자체를 약한 보장 관심으로 둔다.
        ds = _disease_spans(evidence)
        if ds:
            ds.sort(key=lambda s: (s.level, -s.priority, -s.match_score, -s.length, s.start))
            return ds[0]

    semantic = [s for s in candidates if s.slot != "insurance_category"]
    if semantic:
        candidates = semantic
    candidates.sort(key=lambda s: (s.level, -s.priority, -s.match_score, -s.length, s.start))
    return candidates[0]


def _focus_from_selected(selected: Optional[Span]) -> str:
    if selected is None:
        return "null"
    if selected.slot in {"insurance_category", "insurance_intent"}:
        return "null"
    if selected.slot == "weak_intent" and selected.canonical_value in {"추천", "순위", "좋은", "저렴한", "싼", "커뮤니티탐색", "후기"}:
        return "null"
    return selected.canonical_value or "null"



def _weak_need_override(evidence: List[Span], selected: Optional[Span]) -> Optional[str]:
    """상업/탐색 의도는 coverage/category보다 낮은 LEVEL이지만 customer_need_type 정합성에는 중요하다.

    예: '간병인 보험 가격'은 coverage_focus=간병인사용일당이 잡히더라도 고객 니즈는 보험료확인이다.
    단, Hard OOS/Public/Action은 이 override가 뒤집지 않는다.
    """
    if selected and selected.level <= 3:
        return None
    weak = [s for s in evidence if s.slot == "weak_intent" and s.need_hint]
    if not weak:
        return None
    weak.sort(key=lambda s: (-s.priority, -s.match_score, -s.length, s.start))
    top = weak[0]
    if top.need_hint in {"보험료확인", "상품비교", "견적/상담요청", "상품추천탐색", "가입가능성확인", "가입방법확인", "상품조건확인"}:
        return top.need_hint
    return None

def _need_from_selected(selected: Optional[Span], category: str, evidence: List[Span]) -> str:
    disease_ctx = _disease_context(evidence)
    if disease_ctx == "join":
        return "가입가능성확인"
    if disease_ctx == "benefit":
        return "보장범위확인"
    if disease_ctx == "ambiguous" and selected and selected.slot in {"disease_focus", "coverage_focus"}:
        # 질환명 + 보험만으로는 가입가능성을 확정하지 않는다.
        return "보장범위확인"

    override = _weak_need_override(evidence, selected)
    if override:
        return override

    if selected and selected.need_hint:
        if selected.slot == "policy_product_feature":
            if any(s.slot == "definition_signal" or s.surface in {"뜻", "의미", "정의", "뭐야", "무엇"} for s in evidence):
                return "보험용어/제도탐색"
        return selected.need_hint
    if selected:
        if selected.slot == "hard_oos":
            return "질병정보탐색"
        if selected.slot == "public_non_private":
            return "OOS"
        if selected.slot == "coverage_focus":
            return "보장범위확인"
        if selected.slot in {"policy_product_feature", "underwriting_type"}:
            return "상품조건확인"
        if selected.slot == "brand_name":
            return "브랜드상품확인"
        if selected.slot == "insurance_category":
            return "상품추천탐색"
        if selected.slot == "weak_intent":
            return "상품추천탐색"
    return "상품추천탐색" if category != "null" else "OOS"


def _gate_from_selected(selected: Optional[Span], category: str, need: str) -> str:
    if need == "보험용어/제도탐색":
        return "general"
    if selected and selected.gate_hint in ALLOWED_GATE_TYPES:
        return selected.gate_hint
    if selected and selected.level in {1, 2}:
        return "oos"
    if selected and selected.level in {4, 5}:
        return "detailed"
    if selected and selected.level == 3:
        return "general"
    if category != "null":
        return "general"
    if need in {"질병정보탐색", "OOS"}:
        return "oos"
    return "general"


def _make_trace(evidence: List[Span], selected: Optional[Span], blocked: List[str]) -> str:
    parts: List[str] = []
    selected_key = None
    if selected:
        selected_key = (selected.start, selected.end, selected.slot, selected.canonical_value, selected.match_type)
    for s in evidence:
        flag = "true" if selected_key == (s.start, s.end, s.slot, s.canonical_value, s.match_type) else "false"
        parts.append(f"{s.slot}:{s.surface}→{s.canonical_value}|level=L{s.level}|priority={s.priority}|match={s.match_type}|score={s.match_score:.2f}|selected={flag}")
    parts.extend(blocked)
    return "; ".join(parts)


def _model_conflict(need: str, model_hint: str) -> bool:
    if not model_hint or "need=" not in model_hint:
        return False
    m = re.search(r"need=([^|;]+)\|([0-9.]+)", model_hint)
    if not m:
        return False
    try:
        prob = float(m.group(2))
    except Exception:
        prob = 0.0
    return m.group(1).strip() != need and prob >= 0.65


def _confidence_and_review(selected: Optional[Span], category: str, need: str, focus: str, gate: str, model_hint: str, conflict_reason: Optional[str], evidence: Optional[List[Span]] = None) -> Tuple[str, str]:
    evidence = evidence or []
    if _disease_context(evidence) == "ambiguous":
        return "needs_review", "Y"
    if conflict_reason or _model_conflict(need, model_hint):
        return "needs_review", "Y"
    if selected is None:
        return "low", "Y"
    if selected.match_type == "typo":
        return "medium", "Y"
    if selected.match_type == "compact_contains":
        return "medium", "N"
    if gate == "detailed" and focus == "null":
        return "needs_review", "Y"
    if category == "null" and need in {"상품추천탐색", "상품비교", "보험료확인", "견적/상담요청"}:
        return "needs_review", "Y"
    if selected.level in {1, 2, 3, 4}:
        return "high", "N"
    if selected.level == 5:
        return ("high", "N") if category != "null" else ("medium", "N")
    if selected.level in {6, 7, 8}:
        if selected.level == 8 and category in {"null", "기타보험"}:
            return "low", "Y"
        return "medium", "N"
    return "low", "Y"



def _customer_need_detail(need: str, selected: Optional[Span], category: str, evidence: List[Span]) -> str:
    """대분류 customer_need_type을 깨지 않고 세부 니즈를 추가한다.

    예: 보장범위확인 → 질환진단비보장확인 / 수술비보장확인 / 암치료비보장확인
        가입가능성확인 → 유병력가입가능성 / 인수완화조건확인
    """
    ctx = _disease_context(evidence)
    text = _joined_text(evidence)
    focus = selected.canonical_value if selected else ""
    if need == "가입가능성확인":
        if "나이" in text or "연령" in text or "가입연령" in text:
            return "가입연령조건확인"
        if "간편" in text or "고지" in text:
            return "간편고지/심사조건확인"
        if "인수" in text or "완화" in text:
            return "인수완화조건확인"
        if ctx == "join":
            return "유병력가입가능성확인"
        return "가입가능성확인"
    if need == "보장범위확인":
        # 세부 니즈는 전체 trace보다 최종 선택된 focus와 category를 먼저 본다.
        # 치아/치과 계열은 focus에 '치아'가 없어도(스플린트, 지르코니아, 실란트) 치아치료 보장으로 본다.
        if category in {"치아보험", "치과보험", "치아보존보험", "치아보철보험", "치아교정보험", "치과치료보험"}:
            return "치아치료보장확인"
        # 간병인사용일당은 '일당'을 포함하지만 실무적으로는 입원일당보다 간병/요양 보장으로 보는 것이 맞다.
        if "간병" in focus or "요양" in focus or "간호" in focus:
            return "간병/요양보장확인"
        if "수술" in focus:
            return "수술비보장확인"
        if "입원" in focus or "일당" in focus:
            return "입원/일당보장확인"
        if "치료비" in focus or ("치료" in focus and "진단" not in focus):
            return "치료비보장확인"
        if any(k in focus for k in ["진단비", "진단금"]):
            disease_words = {"당뇨", "고혈압", "고지혈증"}
            if any(dw in text or dw in focus for dw in disease_words):
                return "질환진단비보장확인"
            return "진단비보장확인"
        if any(k in text for k in ["진단비", "진단금"]):
            disease_words = {"당뇨", "고혈압", "고지혈증"}
            if any(dw in text or dw in focus for dw in disease_words):
                return "질환진단비보장확인"
            return "진단비보장확인"
        if "수술" in text:
            return "수술비보장확인"
        if "입원" in text or "일당" in text:
            return "입원/일당보장확인"
        if "치료비" in text or "치료" in text:
            return "치료비보장확인"
        if "간병" in text or "요양" in text:
            return "간병/요양보장확인"
        if any(k in text for k in ["임플란트", "틀니", "크라운", "치아", "치과"]):
            return "치아치료보장확인"
        if "암" in text:
            return "암보장확인"
        if focus:
            return f"{focus}확인"
        return "보장범위확인"
    if need == "상품조건확인":
        if selected and selected.slot == "underwriting_type":
            return "인수/간편심사조건확인"
        if selected and selected.canonical_value in {"면책기간", "감액기간", "보장개시일", "책임개시일"}:
            return "약관조건확인"
        if selected and selected.canonical_value in {"체증형", "비갱신형", "갱신형", "무해지형", "저해지형"}:
            return f"{selected.canonical_value}조건확인"
        if selected and selected.canonical_value in {"가입연령", "보험나이"}:
            return "가입연령조건확인"
        return "상품조건확인"
    if need == "청구/보험금문의":
        return "보험금청구방법확인"
    if need == "서류/증빙확인":
        return "청구서류/증빙확인"
    if need == "해지/환급문의":
        return "해지환급금확인"
    if need == "계약관리문의":
        return "계약조회/변경확인"
    if need == "갱신/유지문의":
        return "갱신/유지조건확인"
    if need == "보험료확인":
        return "보험료/가격확인"
    if need == "상품비교":
        if any(k in text for k in ["장단점", "문제점", "차이"]):
            return "장단점/차이확인"
        return "상품비교/순위확인"
    if need == "브랜드상품확인":
        return "보험사/브랜드상품확인"
    if need == "상품추천탐색":
        if any(k in text for k in ["후기", "디시", "필요성", "필요한가", "추천", "문제점"]):
            return "후기/추천/필요성탐색"
        return "일반상품탐색"
    return need

def validate_contract(result: Dict[str, str], evidence: List[Span]) -> Dict[str, str]:
    gate = result.get("gate_type", "general")
    category = result.get("insurance_category", "null")
    need = result.get("customer_need_type", "OOS")
    focus = result.get("evidence_focus", "null")
    if gate not in ALLOWED_GATE_TYPES:
        gate = "general"
    if category not in ALLOWED_CATEGORIES:
        category = "기타보험" if gate != "oos" else "null"
    if need not in ALLOWED_NEEDS:
        need = "OOS" if gate == "oos" else "상품추천탐색"
    if gate == "oos":
        category = "null"
    if focus in {"", "None", "nan", "NaN"}:
        focus = "null"
    allowed_focus = {"null"} | {s.canonical_value for s in evidence}
    if focus not in allowed_focus and result.get("evidence_trace") != "exact_override:selected=true":
        focus = "null"
        if gate == "detailed":
            result["confidence_flag"] = "needs_review"
            result["review_flag"] = "Y"
    result.update({"gate_type": gate, "insurance_category": category, "customer_need_type": need, "evidence_focus": focus})
    return result


def resolve_query(query: str, rulebook: RuleBook, model_hint: str = "", enable_semantic_hints: bool = False) -> Dict[str, str]:
    query_norm = normalize_query(query)
    exact = find_exact_override(query_norm, rulebook)
    if exact:
        return {"query": query, "query_norm": query_norm, **exact}

    exact_spans = detect_phrases(query_norm, rulebook)
    similarity_spans, similarity_hints = detect_similarity_candidates(query_norm, rulebook, exact_spans, enable_semantic_hints=enable_semantic_hints)
    spans = exact_spans + similarity_spans
    protected = build_protected_spans(spans)
    evidence, blocked = extract_evidence(query_norm, spans, protected, rulebook)
    selected = _selected_evidence(evidence)

    conflict_reason = None
    has_public = any(s.slot == "public_non_private" for s in evidence)
    has_private_category = any(s.slot == "insurance_category" for s in evidence)
    if has_public and has_private_category:
        conflict_reason = "public_private_conflict"

    if selected and selected.slot == "hard_oos" and _has_private_insurance_signal(evidence):
        non_oos = [s for s in evidence if s.slot != "hard_oos"]
        selected = _selected_evidence(non_oos)
        conflict_reason = "hard_oos_with_insurance_intent"

    category = _best_category(evidence, selected)
    focus = _focus_from_selected(selected)
    need = _need_from_selected(selected, category, evidence)
    gate = _gate_from_selected(selected, category, need)
    combined_hint = "; ".join([p for p in [model_hint, "; ".join(similarity_hints)] if p])
    confidence, review = _confidence_and_review(selected, category, need, focus, gate, combined_hint, conflict_reason, evidence)
    result = {
        "query": query,
        "query_norm": query_norm,
        "gate_type": gate,
        "insurance_category": category,
        "customer_need_type": need,
        "evidence_focus": focus,
        "customer_need_detail": _customer_need_detail(need, selected, category, evidence),
        "evidence_trace": _make_trace(evidence, selected, blocked),
        "confidence_flag": confidence,
        "model_hint": combined_hint,
        "review_flag": review,
    }
    return validate_contract(result, evidence)
